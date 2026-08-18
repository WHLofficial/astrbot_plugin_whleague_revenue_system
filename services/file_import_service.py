"""CSV/XLSX 文件导入解析层：赛程 / 赛果 / 属性。

- 文件 → 归一化行（与文本导入同一格式），服务层复用既有校验/记账逻辑
- 首行自动识别表头（关键字映射列）；无表头按位置取列
- xlsx 用 openpyxl 只读流式读首 sheet；CSV 自动探测编码（utf-8-sig → gbk → utf-8）
- 单元格统一转字符串，防 Excel 数值精度变形
- 纯解析层，无 DB 依赖；耗时解析走 asyncio.to_thread
"""

import asyncio
import csv
import re
from pathlib import Path

from astrbot.api import logger

_ALLOWED_EXTS = (".csv", ".xlsx")

_HEADER_KEYWORDS = {
    "round": ("轮次", "round", "轮"),
    "week": ("周", "周次", "week", "w"),
    "day": ("天", "日", "day", "d"),
    "weekday": ("星期", "weekday"),
    "time": ("时间", "开赛时间", "time"),
    "home": ("主队", "主场", "home", "主"),
    "away": ("客队", "客场", "away", "客"),
    "team": ("队名", "球队", "team", "队伍", "名称"),
    "influence": ("影响力", "influence", "实力", "strength"),
    "capacity": ("容量", "座位", "capacity", "cap", "座"),
    "tier": ("等级", "级别", "tier", "level", "lv", "grade"),
    "result": ("赛果", "结果", "result", "res", "胜平负"),
    "score": ("比分", "score"),
}

_RESULT_ALIASES = {
    "胜": "W", "w": "W", "win": "W",
    "平": "D", "d": "D", "draw": "D",
    "负": "L", "l": "L", "loss": "L",
}


class FileImportError(Exception):
    pass


def _cell_to_str(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        if value == int(value):
            return str(int(value))
        return format(value, ".15f").rstrip("0").rstrip(".")
    if isinstance(value, int):
        return str(value)
    return str(value).strip()


def _read_csv(path: Path) -> list[list[str]]:
    raw = path.read_bytes()
    text = None
    for encoding in ("utf-8-sig", "gbk", "utf-8"):
        try:
            text = raw.decode(encoding)
            break
        except (UnicodeDecodeError, LookupError):
            continue
    if text is None:
        text = raw.decode("utf-8", errors="replace")
    lines = text.splitlines()
    # Excel 导出的制表符分隔（TSV）优先按 tab 切分，否则按标准 CSV（逗号）
    sample = next((l for l in lines if l.strip()), "")
    if "\t" in sample:
        return [[c.strip() for c in l.split("\t")] for l in lines]
    return [[c.strip() for c in row] for row in csv.reader(lines)]


def _read_xlsx(path: Path) -> list[list[str]]:
    from openpyxl import load_workbook

    rows = []
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.worksheets[0]
        for row in ws.iter_rows(values_only=True):
            rows.append([_cell_to_str(v) for v in row])
    finally:
        wb.close()
    return rows


def read_rows(path, max_size_mb: float = 20.0) -> list[list[str]]:
    """读取文件为行列字符串；校验存在性、扩展名与大小上限。"""
    p = Path(path)
    if not p.is_file():
        raise FileImportError(f"文件不存在: {p.name}")
    if p.suffix.lower() not in _ALLOWED_EXTS:
        raise FileImportError("仅支持 .csv / .xlsx 文件")
    try:
        size_mb = p.stat().st_size / (1024 * 1024)
    except OSError:
        raise FileImportError("无法读取文件信息")
    if size_mb > max_size_mb:
        raise FileImportError(f"文件超过大小上限（{max_size_mb:g} MB）")
    try:
        if p.suffix.lower() == ".csv":
            return _read_csv(p)
        return _read_xlsx(p)
    except FileImportError:
        raise
    except Exception as e:
        logger.error(f"File parse error: {e}")
        raise FileImportError(f"文件解析失败: {e}")


def _non_empty_rows(rows: list[list[str]]) -> list[list[str]]:
    return [r for r in rows if any(str(c).strip() for c in r)]


def _detect_header(rows: list[list[str]]) -> tuple[list[list[str]], dict | None]:
    """首行命中任一表头关键字则按列名映射；否则返回 None（按位置取列）。"""
    rows = _non_empty_rows(rows)
    if not rows:
        return [], None
    first = [str(c).strip().lower() for c in rows[0]]
    col_map: dict = {}
    matched = False
    for kind, keywords in _HEADER_KEYWORDS.items():
        for i, cell in enumerate(first):
            if cell in keywords:
                col_map[kind] = i
                matched = True
                break
    if not matched:
        return rows, None
    return rows[1:], col_map


def _cell(row: list[str], col_map: dict | None, kind: str,
          fallback_idx: int | None) -> str:
    if col_map and kind in col_map:
        idx = col_map[kind]
        return str(row[idx]).strip() if idx < len(row) else ""
    if fallback_idx is not None and fallback_idx < len(row):
        return str(row[fallback_idx]).strip()
    return ""


def build_fixture_lines(rows: list[list[str]]) -> tuple[list[str], list[str]]:
    """归一化为「轮次 主队 客队 [周] [天] [时间]」文本行。返回 (lines, errors)。"""
    from . import formula

    data, col_map = _detect_header(rows)
    lines, errors = [], []
    start = 2 if col_map else 1
    for i, row in enumerate(data):
        rownum = i + start
        round_raw = _cell(row, col_map, "round", 0 if not col_map else None)
        home = _cell(row, col_map, "home", 1 if not col_map else None)
        away = _cell(row, col_map, "away", 2 if not col_map else None)
        week_raw = _cell(row, col_map, "week", 3 if not col_map else None)
        day_raw = _cell(row, col_map, "day", 4 if not col_map else None)
        time_raw = _cell(row, col_map, "time", 5 if not col_map else None)
        if not home or not away:
            if all(not str(c).strip() for c in row):
                continue  # 整行空白
            errors.append(f"第{rownum}行: 主客队缺失，需为「轮次 主队 客队 [周] [天] [时间]」")
            continue
        if home == away:
            errors.append(f"第{rownum}行: 主客队不能相同")
            continue
        round_token = round_raw if round_raw else "1"
        # 纯文字轮次（无数字）允许透传，由服务层按「同名即为同一轮」登记并改写
        digits = re.sub(r"\D", "", round_token)
        if digits and int(digits) < 1:
            errors.append(f"第{rownum}行: 轮次需为正数「{round_token}」")
            continue
        ok = True
        week = day = time = None
        if week_raw:
            m = re.match(r"^[Ww]?(\d+)$", week_raw)
            if not m or int(m.group(1)) < 1:
                errors.append(f"第{rownum}行: 周需为 W+数字 或数字「{week_raw}」")
                ok = False
            else:
                week = int(m.group(1))
        if day_raw:
            m = re.match(r"^[Dd]?(\d+)$", day_raw)
            if not m or not (1 <= int(m.group(1)) <= 7):
                errors.append(f"第{rownum}行: 天需为 1-7「{day_raw}」")
                ok = False
            else:
                day = int(m.group(1))
        if time_raw:
            try:
                time = formula.norm_time(time_raw)
            except ValueError as e:
                errors.append(f"第{rownum}行: {e}")
                ok = False
        if not ok:
            continue
        parts = [round_token, home, away]
        if week is not None:
            parts.append(f"W{week}")
        if day is not None:
            parts.append(f"D{day}")
        if time is not None:
            parts.append(time)
        lines.append(" ".join(parts))
    return lines, errors


def build_result_lines(rows: list[list[str]]) -> tuple[list[str], list[str]]:
    """归一化为「主队 胜/平/负 [比分]」文本行。返回 (lines, errors)。

    赛果列可显式填 胜/平/负，也可空缺或直接写比分（如 2-1 / 0-0PK2-4），
    由比分自动推导主队胜平负。
    """
    from . import formula

    data, col_map = _detect_header(rows)
    lines, errors = [], []
    start = 2 if col_map else 1
    for i, row in enumerate(data):
        rownum = i + start
        if col_map:
            home = _cell(row, col_map, "home", None) or _cell(row, col_map, "team", None)
            res_raw = _cell(row, col_map, "result", None)
            score_raw = _cell(row, col_map, "score", None)
        else:
            home = str(row[0]).strip() if row else ""
            res_raw = str(row[1]).strip() if len(row) >= 2 else ""
            score_raw = " ".join(str(c).strip() for c in row[2:] if str(c).strip()).strip() or None
        if not home:
            if all(not str(c).strip() for c in row):
                continue  # 整行空白
            errors.append(f"第{rownum}行: 主队缺失")
            continue
        res = res_raw.strip()
        result = _RESULT_ALIASES.get(res.lower())
        score_txt = score_raw
        if result is None:
            # 赛果列空缺或本身就是比分 → 由比分列/赛果列推导
            result = formula.result_from_score(score_txt or res)
            if result is not None and not score_txt and res:
                score_txt = res
            if result is None:
                errors.append(
                    f"第{rownum}行: 赛果需为 胜/平/负（W/D/L）或比分（如 2-1 / 0-0PK2-4）「{res_raw}」"
                )
                continue
        line = f"{home} {result}"
        if score_txt:
            line += f" {score_txt}"
        lines.append(line)
    return lines, errors


def build_attribute_rows(rows: list[list[str]]) -> tuple[list, list[str]]:
    """归一化为 (队名, 影响力, 容量, 等级) 记录。返回 (records, errors)。"""
    data, col_map = _detect_header(rows)
    records, errors = [], []
    start = 2 if col_map else 1
    for i, row in enumerate(data):
        rownum = i + start
        team = _cell(row, col_map, "team", 0 if not col_map else None)
        if not team:
            if all(not str(c).strip() for c in row):
                continue  # 整行空白
            errors.append(f"第{rownum}行: 队名缺失")
            continue
        influence_raw = _cell(row, col_map, "influence", 1 if not col_map else None)
        capacity_raw = _cell(row, col_map, "capacity", 2 if not col_map else None)
        tier_raw = _cell(row, col_map, "tier", 3 if not col_map else None)

        influence = capacity = tier = None
        try:
            if influence_raw and influence_raw != "-":
                influence = float(influence_raw)
            if capacity_raw and capacity_raw != "-":
                capacity = int(float(capacity_raw))
            if tier_raw and tier_raw != "-":
                tier = int(float(tier_raw))
        except ValueError:
            errors.append(f"第{rownum}行: 属性格式错误（影响力/容量/等级）")
            continue
        records.append((team, influence, capacity, tier))
    return records, errors


# ─── 异步入口（耗时解析进线程池） ─────────────────────────

def _max_size_mb(cfg) -> float:
    try:
        return float(cfg.get("import_max_file_size_mb", 20.0))
    except (TypeError, ValueError):
        return 20.0


async def parse_fixture_file(cfg, path) -> dict:
    """读取文件并归一化为赛程文本行。返回 {"lines", "errors"}。"""
    rows = await asyncio.to_thread(read_rows, path, _max_size_mb(cfg))
    lines, errors = await asyncio.to_thread(build_fixture_lines, rows)
    return {"lines": lines, "errors": errors}


async def parse_result_file(cfg, path) -> dict:
    """读取文件并归一化为赛果文本行。返回 {"lines", "errors"}。"""
    rows = await asyncio.to_thread(read_rows, path, _max_size_mb(cfg))
    lines, errors = await asyncio.to_thread(build_result_lines, rows)
    return {"lines": lines, "errors": errors}


async def parse_attribute_file(cfg, path) -> dict:
    """读取文件并归一化为属性记录。返回 {"records", "errors"}。"""
    rows = await asyncio.to_thread(read_rows, path, _max_size_mb(cfg))
    records, errors = await asyncio.to_thread(build_attribute_rows, rows)
    return {"records": records, "errors": errors}


def cleanup_file(path) -> None:
    """删除解析后的临时文件（AstrBot 下载的附件），失败静默。"""
    if not path:
        return
    try:
        p = Path(path)
        if p.is_file():
            p.unlink()
    except OSError:
        pass